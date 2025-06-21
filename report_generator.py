import datetime
import pandas as pd
import numpy as np
import requests
import pickle
import re
from io import StringIO
from datetime import datetime
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

################################################################
# 원하는 행 추출 함수의 헬퍼 함수: 조건에 따라 행 찾기
def _find_row(df, key_column, acc, exact_match):
    if exact_match:
        return df[df[key_column] == acc]
    else:
        return df[df[key_column].str.contains(acc, na=False)]

# 첫 행 추출 함수
def extract_first_row(df, index_column, target_column):
    i_df = pd.DataFrame()

    # 행이 존재하면 첫 번째 행 추출
    if not df.empty:
        first_row = df.iloc[[0]][[index_column, target_column]]  # iloc[[0]]은 DataFrame 유지
        i_df = pd.concat([i_df, first_row], ignore_index=True)
    else:
        # 빈 데이터프레임일 경우, 오류 대신 표시용 값 반환
        i_df = pd.DataFrame({
            index_column: ["(empty)"],
            target_column: ["len(df)=0"]
        })

    return i_df

# 원하는 행 추출 함수
def extract_data(df, index_column, key_column, target_column, 
                 account_list, exact_match_acc=True, 
                 backup_accounts=None, exact_match_backup=True,
                 fallback_accounts=None, fallback_key_column=None, exact_match_fallback=True, ascending=True) :
    
    # i_df 설정
    i_df = pd.DataFrame()

    # log_df 저장할 list
    log_list = []

    # account_list가 str형태면 리스트로 만들어주기
    if isinstance(account_list, str):
        account_list = [account_list]
    
    # 1. account_list 검색
    for acc in account_list:
        s_row = _find_row(df, key_column, acc, exact_match_acc)

        # 2. 1에서 못찾고, backup_accounts가 존재하고, backup_accounts의 key중에 acc가 있다면
        if s_row.empty and backup_accounts and acc in backup_accounts:
            # backup_accounts의 value에 있는 list들에 대해서 s_row 파싱
            for backup_acc in backup_accounts[acc]:
                s_row = _find_row(df, key_column, backup_acc, exact_match_backup)
                if not s_row.empty:
                    break  # 찾았으면 멈춤
                
        # 3. 2에서도 못찾고, fallback_accounts 존재하고, fallback_accounts의 key중에 acc가 있다면
        if s_row.empty and fallback_accounts and acc in fallback_accounts:
            # fallback_accounts의 value에 있는 list들에 대해서 s_row 파싱
            for fallback_acc in fallback_accounts[acc]:
                # account_nm(한글 칼럼으로)으로 파싱하면 exact_match=True여도 여러개의 행이 나올 수 있다.  
                # 그래서 target_column의 값 기준으로 가장 작은값이 위로 정렬
                s_row = _find_row(df, fallback_key_column, fallback_acc, exact_match_fallback)

                if not s_row.empty :
                    break  # 찾았으면 멈춤

        # s_row가 2행 이상이면 
        if s_row.shape[0] > 1 :
            s_row_over_2 = s_row
            
            # 2개 이상 데이터에서 크거나 작은 값을 가려내기 위해 숫자형으로 바꾸고 sort_values
            s_row_over_2.loc[:, target_column] = pd.to_numeric(df.loc[:, target_column].replace(',', '', regex=True), errors='coerce')
            s_row = s_row_over_2.sort_values(by=target_column, ascending=ascending)

            # log 저장
            message = "s_row 2행 이상"
            log_df = capture_log(extract_data, message, s_row_over_2)
            log_list.append(log_df)

        s_row.reset_index(drop=True, inplace=True)
        
        if not s_row.empty:
            i_df = pd.concat([i_df, s_row.loc[[0]]], ignore_index=True)[[index_column, target_column]]
        else:
            i_df = pd.concat(
                [i_df, pd.DataFrame({index_column: [acc], target_column: ["len(s_row) = 0"]})],
                ignore_index=True
            )
    return i_df, log_list
################################################################
# 문자와 숫자가 섞여있을 때 숫자만 가져오는 함수
def extract_number(s):
    return int(re.sub(r'[^\d]', '', s))

# 문자를 숫자로 바꾸는 함수
def str_to_num(col):
    return pd.to_numeric(col.replace(',', '', regex=True), errors='coerce')

# pickle에서 report_data 가져오는 함수
def load_report_data(save_path):
    with open(save_path, 'rb') as f:
        report_data = pickle.load(f)

    return report_data

# excel 열 자동조절 함수
def auto_adjust_column_width(ws, max_width=30):
    for col in ws.columns:
        max_len = max((len(str(cell.value)) for cell in col if cell.value), default=0)
        adjusted_width = min(max_len + 2, max_width)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = adjusted_width

# log_list 저장하는 함수 
def capture_log(func_name, message, value_df):

    log_df = value_df.copy()
    log_df["log"] = message
    log_df["출처"] = func_name

    return log_df
################################################################
# conpany_info 저장하는 함수
def parse_company_info(company_dict):
    exchange_map = {'Y': 'KRX', 'K': 'KOSDAQ'}
    exchange = exchange_map.get(company_dict.get("corp_cls"), "없음")
    est_date_raw = company_dict.get("est_dt", "")
    try:
        est_date = datetime.strptime(est_date_raw, "%Y%m%d").strftime("%Y-%m-%d")
    except ValueError:
        est_date = "날짜 오류"
    data = [
        ("종목명", company_dict["stock_name"]),
        ("종목코드", company_dict["stock_code"]),
        ("거래소", exchange),
        ("설립일", est_date),
        ("본사주소", company_dict["adres"]),
        ("CEO", company_dict["ceo_nm"])
    ]
    return pd.DataFrame(data, columns=["항목", "내용"])

# additional_company_info 저장하는 함수
def fetch_additional_company_info(s_code):
    url_tmpl_2 = 'http://companyinfo.stock.naver.com/v1/company/c1020001.aspx?cmp_cd=%s'
    url_2 = url_tmpl_2 % s_code
    html_text_2 = requests.get(url_2).text

    dfs = pd.read_html(StringIO(html_text_2))

    df2 = dfs[1]
    data2 = [
        (df2.loc[2, "항목"], df2.loc[2, "내용"]),
        (df2.loc[4, "항목"], df2.loc[4, "내용"]),
        (df2.loc[5, "항목.1"], df2.loc[5, "내용.1"]),
        (df2.loc[1, "항목"], df2.loc[1, "내용"])
    ]

    cov = pd.DataFrame(data2, columns=["항목", "내용"])

    product = dfs[3]
    capital = dfs[8]
    related = dfs[9]
    affiliate = dfs[10]

    return cov, product, capital, related, affiliate
################################################################
# 연간 손익계산서 전역변수
ACCOUNT_LISTS_IS = {
    # a기간
    'a': [
        "ifrs_OperatingExpense",
        "ifrs_CostOfSales", 
        "dart_TotalSellingGeneralAdministrativeExpenses", 
        "dart_OperatingIncomeLoss",
        "ifrs_ComprehensiveIncome", 
        "ifrs_ComprehensiveIncomeAttributableToOwnersOfParent"
        ],

    # b기간
    'b': [
        "ifrs-full_OperatingExpense",
        "ifrs-full_CostOfSales", 
        "dart_TotalSellingGeneralAdministrativeExpenses", 
        "dart_OperatingIncomeLoss",
        "ifrs-full_ComprehensiveIncome", 
        "ifrs-full_ComprehensiveIncomeAttributableToOwnersOfParent"
        ],

    # c기간
    'c': [
        "ifrs-full_Revenue",
        "ifrs-full_OperatingExpense",
        "ifrs-full_CostOfSales", 
        "dart_TotalSellingGeneralAdministrativeExpenses", 
        "dart_OperatingIncomeLoss",
        "ifrs-full_ComprehensiveIncome", 
        "ifrs-full_ComprehensiveIncomeAttributableToOwnersOfParent"
    ],
    # 보험업
    'insurance': [  
        "dart_OperatingIncomeInsurance",
        "ifrs-full_InvestmentIncome",
        "dart_OtherOperatingIncome",
        "dart_OperatingExpenseInsurance",
        "dart_OperatingExpenseInvestment",
        "dart_OtherOperatingExpense",
        "ifrs-full_ProfitLossFromOperatingActivities",
        "ifrs-full_ComprehensiveIncome",
        "ifrs-full_ComprehensiveIncomeAttributableToOwnersOfParent"
    ]
}

# 연간 손익계산서 BACKUP 전역변수 
BACKUP_ACCOUNTS_IS = {
    'a': {
        "ifrs_OperatingExpense": ["ifrs-full_OperatingExpense", "dart_OperatingExpenses"],
        "ifrs_CostOfSales": ["ifrs-full_CostOfSales"],
        "ifrs_ComprehensiveIncome": ["ifrs-full_ComprehensiveIncome"],
        "ifrs_ComprehensiveIncomeAttributableToOwnersOfParent": ["ifrs-full_ComprehensiveIncomeAttributableToOwnersOfParent"]
    },
    
    'b': {
        "ifrs-full_OperatingExpense" : ["dart_OperatingExpenses"]
    },
    'c': {
                
        "dart_TotalSellingGeneralAdministrativeExpenses": ["ifrs-full_SellingGeneralAndAdministrativeExpense"],
    },

}

# 연간 손익계산서 FALLBACK 전역변수 
FALLBACK_ACCOUNTS_IS = {
    'a': {
        "ifrs_ComprehensiveIncomeAttributableToOwnersOfParent": ["지배기업 소유주지분", 
                                                                 "지배기업의 소유주에 귀속될 총포괄이익",
                                                                 "지배기업의 소유주에게 귀속되는 총포괄손익",
                                                                 "지배기업소유주지분"]
    },

    'b': {
        "ifrs-full_ComprehensiveIncomeAttributableToOwnersOfParent": ["지배기업 소유주지분",
                                                                      "지배기업소유주지분",
                                                                      "지배기업의 소유주에 귀속될 총포괄이익"]
    },

    'c': {
        
        "ifrs-full_Revenue": ["영업수익", "매출액"],
        "ifrs-full_ComprehensiveIncomeAttributableToOwnersOfParent" : ["지배기업소유주지분"]

    }

}

# 연간 손익계산서 parse
def parse_annual_income_statements(fs_list, industry):

    # 연간 손익계산서
    i_s = pd.DataFrame()

    # log_list
    log_list = []

    for df in fs_list:

        # concat할 i_df 설정
        i_df = pd.DataFrame()

        # 연도 및 필요한 원본 데이터 설정
        year = int(df.columns.name)  
        df = df[df['sj_div'].isin(['IS', 'CIS'])][['account_id', 'account_nm', 'thstrm_amount']]

        # 연결재무제표와 별도재무제표 구분
        is_consolidated = df.index.name == "연결재무제표"
        suffix = "(연결)" if is_consolidated else "(별도)"

        # i_df의 index_column, 파싱할 key_column, 가져올 target_column 설정
        index_column = 'account_nm'
        key_column = "account_id"
        target_column = 'thstrm_amount'

        # 2015 ~ 2018년 처리 로직
        if year <= 2018:
            # a기간 으로 설정
            period = "a"

            # 각 기간에 해당하는 acount_list, backup_accounts, fallback_key_column 설정
            account_list = ACCOUNT_LISTS_IS[period]
            backup_accounts = BACKUP_ACCOUNTS_IS.get(period, {}) 
            fallback_accounts = FALLBACK_ACCOUNTS_IS.get(period, {}) 
            fallback_key_column = "account_nm"

            # i_df 가져오기
            # 1. i_s의 0행인 매출은 df의 0행에서 가져오기
            first_row = df.iloc[[0]][[index_column, target_column]]
            i_df = pd.concat([i_df, first_row], ignore_index=True)

            # 2. i_s의 1행 이하 가져오기
            # add_df, add_list 설정
            add_df, add_list = extract_data(df, index_column, key_column, target_column, 
                                account_list, exact_match_acc=True, 
                                backup_accounts=backup_accounts, exact_match_backup=True,
                                fallback_accounts=fallback_accounts, fallback_key_column=fallback_key_column, exact_match_fallback=True)  

            # i_df, add_df concat
            i_df = pd.concat([i_df,
                              add_df], 
                              ignore_index=True
                              )
            
            # add_list 안의 df들에 index.name 설정
            for df in add_list :
                df_name = df.index.name
                df.index.name = f"{df_name} {year}"

            # lost_list에 하나씩 extend
            log_list.extend(add_list)

        # 2019 ~ 2022년 처리 로직
        elif 2019 <= year <= 2022:
            # b기간 으로 설정
            period = "b"

            # 각 기간에 해당하는 acount_list, backup_accounts, fallback_key_column 설정
            account_list = ACCOUNT_LISTS_IS[period]
            backup_accounts = BACKUP_ACCOUNTS_IS.get(period, {}) 
            fallback_accounts = FALLBACK_ACCOUNTS_IS.get(period, {}) 
            fallback_key_column = "account_nm"

            # i_s의 0행인 매출은 df의 0행에서 가져오기
            first_row = df.iloc[[0]][[index_column, target_column]]
            i_df = pd.concat([i_df, first_row], ignore_index=True)
            
            # i_s의 1행 이하 가져오기
            # add_df, add_list 설정
            add_df, add_list = extract_data(df, index_column, key_column, target_column, 
                                account_list, exact_match_acc=True, 
                                backup_accounts=backup_accounts, exact_match_backup=True,
                                fallback_accounts=fallback_accounts, fallback_key_column=fallback_key_column, exact_match_fallback=True)

            # i_df, add_df concat
            i_df = pd.concat([i_df,
                              add_df], 
                              ignore_index=True
                              )
            
            # add_list 안의 df들에 index.name 설정
            for df in add_list :
                df_name = df.index.name
                df.index.name = f"{df_name} {year}"

            # lost_list에 하나씩 extend
            log_list.extend(add_list)

        # 이상 다른 경우
        else:
            # 보험업일 경우
            if industry == "insurance" :
                # peroid를 "insurance"로 설정
                period = "insurance"

                # 각 기간에 해당하는 acount_list, backup_accounts, fallback_key_column 설정
                account_list = ACCOUNT_LISTS_IS[period]
                # backup_accounts = BACKUP_ACCOUNTS_IS.get(period, {}) 
                # fallback_accounts = FALLBACK_ACCOUNTS_IS.get(period, {}) 
                # fallback_key_column = "account_nm"

                # add_df, add_list 설정
                add_df, add_list = extract_data(df, index_column, key_column, target_column, 
                                    account_list, exact_match_acc=True)
                
                # i_df, add_df concat
                i_df = pd.concat([i_df,
                                add_df], 
                                ignore_index=True
                                )
                
                # add_list 안의 df들에 index.name 설정
                for df in add_list :
                    df_name = df.index.name
                    df.index.name = f"{df_name} {year}"

                # lost_list에 하나씩 extend
                log_list.extend(add_list)

            # 2023년 이상 처리 로직
            else :
                # c기간 으로 설정
                period = "c"

                # 각 기간에 해당하는 acount_list, backup_accounts, fallback_key_column 설정
                account_list = ACCOUNT_LISTS_IS[period]
                backup_accounts = BACKUP_ACCOUNTS_IS.get(period, {}) 
                fallback_accounts = FALLBACK_ACCOUNTS_IS.get(period, {}) 
                fallback_key_column = "account_nm"

                # add_df, add_list 설정
                add_df, add_list = extract_data(df, index_column, key_column, target_column, 
                                    account_list, exact_match_acc=True, 
                                    backup_accounts=backup_accounts, exact_match_backup=True,
                                    fallback_accounts=fallback_accounts, fallback_key_column=fallback_key_column, exact_match_fallback=True)
                
                # i_df, add_df concat
                i_df = pd.concat([i_df,
                                add_df], 
                                ignore_index=True
                                )
                
                # add_list 안의 df들에 index.name 설정
                for df in add_list :
                    df_name = df.index.name
                    df.index.name = f"{df_name} {year}"

                # lost_list에 하나씩 extend
                log_list.extend(add_list)
                
        i_df.rename(columns={target_column: f"{year}{suffix}"}, inplace=True)

        i_s = pd.concat([i_s, i_df], axis=1)

    i_s.set_index('account_nm', inplace=True)

    # 모든 값이 "len(s_row) = 0"인 행 제거
    i_s = i_s[~(i_s == "len(s_row) = 0").all(axis=1)]

    return i_s, log_list

# 분기 손익계산서 parse
def parse_quarterly_income_statements(qfs_list, industry):

    # 분기 손익계산서
    q_s = pd.DataFrame()
    
    # log_list 설정
    log_list = []

    for df in qfs_list:

        # concat할 i_df 설정
        i_df = pd.DataFrame()

        # 연도, 분기, 필요한 원본데이터 설정
        label = df.columns.name  # e.g., "2019 3분기"
        year = int(label.split()[0])
        quarter = int(label.split()[1].strip().replace("분기", ""))
        df = df[df['sj_div'].isin(['IS', 'CIS'])][['account_id', 'account_nm', 'thstrm_amount']]

        # 연결재무제표와 별도재무제표 구분
        is_consolidated = df.index.name == "연결재무제표"
        suffix = "(연결)" if is_consolidated else "(별도)"

        # i_df의 index_column, 파싱할 key_column, 가져올 target_column 설정
        index_column = 'account_nm'
        key_column = "account_id"
        target_column = 'thstrm_amount'

        # 19년 2분기까지 
        if (year < 2019) or (year == 2019 and quarter <= 2) :
            
            # a기간 으로 설정
            period = "a"

            # 각 기간에 해당하는 acount_list, backup_accounts, fallback_key_column 설정
            account_list = ACCOUNT_LISTS_IS[period]
            backup_accounts = BACKUP_ACCOUNTS_IS.get(period, {}) 
            fallback_accounts = FALLBACK_ACCOUNTS_IS.get(period, {}) 
            fallback_key_column = "account_nm"

            # i_df 가져오기
            # 1. i_s의 0행인 매출은 df의 0행에서 가져오기
            first_row = df.iloc[[0]][[index_column, target_column]]
            i_df = pd.concat([i_df, first_row], ignore_index=True)
            
            # 2. i_s의 1행 이하 가져오기
            # add_df, add_list 설정
            add_df, add_list = extract_data(df, index_column, key_column, target_column, 
                                account_list, exact_match_acc=True, 
                                backup_accounts=backup_accounts, exact_match_backup=True,
                                fallback_accounts=fallback_accounts, fallback_key_column=fallback_key_column, exact_match_fallback=True)
            
            # i_df, add_df concat
            i_df = pd.concat([i_df,
                            add_df], 
                            ignore_index=True
                            )

            # add_list 안의 df들에 index.name 설정
            for df in add_list :
                df_name = df.index.name
                df.index.name = f"{df_name} {year}"

            # lost_list에 하나씩 extend
            log_list.extend(add_list)
            
        # 23년 2분기까지 
        elif (year < 2023) or (year == 2023 and quarter <= 2) :
            
            # b기간 으로 설정
            period = "b"

            # 각 기간에 해당하는 acount_list, backup_accounts, fallback_key_column 설정
            account_list = ACCOUNT_LISTS_IS[period]
            backup_accounts = BACKUP_ACCOUNTS_IS.get(period, {}) 
            fallback_accounts = FALLBACK_ACCOUNTS_IS.get(period, {}) 
            fallback_key_column = "account_nm"

            # i_s의 0행인 매출은 df의 0행에서 가져오기
            first_row = df.iloc[[0]][[index_column, target_column]]
            i_df = pd.concat([i_df, first_row], ignore_index=True)
            
            # i_s의 1행 이하 가져오기
            # add_df, add_list 설정
            add_df, add_list = extract_data(df, index_column, key_column, target_column, 
                                account_list, exact_match_acc=True, 
                                backup_accounts=backup_accounts, exact_match_backup=True,
                                fallback_accounts=fallback_accounts, fallback_key_column=fallback_key_column, exact_match_fallback=True)

            # i_df, add_df concat
            i_df = pd.concat([i_df,
                            add_df], 
                            ignore_index=True
                            )
            
            # add_list 안의 df들에 index.name 설정
            for df in add_list :
                df_name = df.index.name
                df.index.name = f"{df_name} {year}"

            # lost_list에 하나씩 extend
            log_list.extend(add_list)
            
        # 이하 다른 경우
        else:
            # 보험업일 경우
            if industry == "insurance" :
                # peroid를 "insurance"로 설정
                period = "insurance"

                # 각 기간에 해당하는 acount_list, backup_accounts, fallback_key_column 설정
                account_list = ACCOUNT_LISTS_IS[period]
                # backup_accounts = BACKUP_ACCOUNTS_IS.get(period, {}) 
                # fallback_accounts = FALLBACK_ACCOUNTS_IS.get(period, {}) 
                # fallback_key_column = "account_nm"

                # i_df 가져오기
                # add_df, add_list 설정
                add_df, add_list = extract_data(df, index_column, key_column, target_column, 
                                    account_list, exact_match_acc=True)
                
                # i_df, add_df concat
                i_df = pd.concat([i_df,
                                add_df], 
                                ignore_index=True
                                )

                # add_list 안의 df들에 index.name 설정
                for df in add_list :
                    df_name = df.index.name
                    df.index.name = f"{df_name} {year}"

                # lost_list에 하나씩 extend
                log_list.extend(add_list) 

            # 2023년 3분기 이상
            else :
                # c기간 으로 설정
                period = "c"
                
                # 각 기간에 해당하는 acount_list, backup_accounts, fallback_key_column 설정
                account_list = ACCOUNT_LISTS_IS[period]
                backup_accounts = BACKUP_ACCOUNTS_IS.get(period, {}) 
                fallback_accounts = FALLBACK_ACCOUNTS_IS.get(period, {}) 
                fallback_key_column = "account_nm"

                # i_df 가져오기
                # add_df, add_list 설정
                add_df, add_list = extract_data(df, index_column, key_column, target_column, 
                                    account_list, exact_match_acc=True, 
                                    backup_accounts=backup_accounts, exact_match_backup=True,
                                    fallback_accounts=fallback_accounts, fallback_key_column=fallback_key_column, exact_match_fallback=True)
                
                # i_df, add_df concat
                i_df = pd.concat([i_df,
                                add_df], 
                                ignore_index=True
                                )
                
                # add_list 안의 df들에 index.name 설정
                for df in add_list :
                    df_name = df.index.name
                    df.index.name = f"{df_name} {year}"

                # lost_list에 하나씩 extend
                log_list.extend(add_list)
                
        i_df.rename(columns={target_column: f"{label}{suffix}"}, inplace=True)

        q_s = pd.concat([q_s, i_df], axis=1)

    q_s.set_index('account_nm', inplace=True)

    # 모든 값이 "len(s_row) = 0"인 행 제거
    q_s = q_s[~(q_s == "len(s_row) = 0").all(axis=1)]

    return q_s, log_list
################################################################
# 연간 재무상태표 전역변수
ACCOUNT_LISTS_BS = {
    'a': [
        "ifrs_CurrentAssets",
        "ifrs_NoncurrentAssets", 
        "ifrs_CurrentLiabilities",
        "ifrs_NoncurrentLiabilities", 
        "ifrs_Equity",
        "ifrs_NoncontrollingInterests",
        "ifrs_EquityAttributableToOwnersOfParent"
        ],

    'b': [
        "ifrs-full_CurrentAssets",
        "ifrs-full_NoncurrentAssets", 
        "ifrs-full_CurrentLiabilities",
        "ifrs-full_NoncurrentLiabilities", 
        "ifrs-full_Equity",
        "ifrs-full_NoncontrollingInterests",
        "ifrs-full_EquityAttributableToOwnersOfParent"
        ],
        
    # 보험업
    'insurance': [
        "ifrs-full_Assets",
        "ifrs-full_CashAndCashEquivalents",
        "ifrs-full_Liabilities",
        "ifrs-full_InsuranceContractsIssuedThatAreLiabilities",
        "ifrs-full_Equity",
        "ifrs-full_EquityAttributableToOwnersOfParent"
        ]
}

# 연간 재무상태표 BACKUP 전역변수
BACKUP_ACCOUNTS_BS = {
    'a': {
        "ifrs_CurrentAssets": ["ifrs-full_CurrentAssets"],
        "ifrs_NoncurrentAssets": ["ifrs-full_NoncurrentAssets"],
        "ifrs_CurrentLiabilities": ["ifrs-full_CurrentLiabilities"],
        "ifrs_NoncurrentLiabilities": ["ifrs-full_NoncurrentLiabilities"],
        "ifrs_Equity": ["ifrs-full_Equity"],
        "ifrs_NoncontrollingInterests": ["ifrs-full_NoncontrollingInterests"],
        "ifrs_EquityAttributableToOwnersOfParent": ["ifrs-full_EquityAttributableToOwnersOfParent"],
    },
    'b': {
        # "ifrs_CurrentAssets": [""],
        # "ifrs_NoncurrentAssets": [""],
        # "ifrs_CurrentLiabilities": [""],
        # "ifrs_NoncurrentLiabilities": [""],
        # "ifrs_Equity": [""],
        # "ifrs_NoncontrollingInterests": [""],
        # "ifrs_EquityAttributableToOwnersOfParent": [""],
    },

    'insurance': {

    }
}

# 연간 재무상태표 FALLBACK 전역변수 
FALLBACK_ACCOUNTS_BS = {
    'a': {
        "ifrs_NoncontrollingInterests": ["비지배지분"]
    },

    'b': {
        "ifrs-full_NoncontrollingInterests": ["비지배지분"]
    }
}

# 연간 재무상태표 parse
def parse_annual_balance_sheets(fs_list, industry):

    # 연간 재무상태표
    b_s = pd.DataFrame()

    for df in fs_list:

        # concat할 i_df 설정
        i_df = pd.DataFrame()

        # log_list 설정
        log_list = []

        # 연도 및 필요한 원본 데이터 설정
        year = int(df.columns.name)  
        df = df[df['sj_div'] == 'BS'][['account_id', 'account_nm', 'thstrm_amount']]
    
        # 연결재무제표와 별도재무제표 구분
        is_consolidated = df.index.name == "연결재무제표"
        suffix = "(연결)" if is_consolidated else "(별도)"

        # i_df의 index_column, 파싱할 key_column, 가져올 target_column 설정
        index_column = 'account_nm'
        key_column = "account_id"
        target_column = 'thstrm_amount'

        # 2015 ~ 2018년 처리 로직
        if year <= 2018:

            # a기간 으로 설정
            period = "a"

            # 각 기간에 해당하는 acount_list, backup_accounts, fallback_key_column 설정
            account_list = ACCOUNT_LISTS_BS[period]
            backup_accounts = BACKUP_ACCOUNTS_BS.get(period, {}) 
            fallback_accounts = FALLBACK_ACCOUNTS_BS.get(period, {})
            fallback_key_column = "account_nm"
            
            # i_df 가져오기
            # add_df, add_list 설정
            add_df, add_list = extract_data(df, index_column, key_column, target_column, 
                                account_list, exact_match_acc=True, 
                                backup_accounts=backup_accounts, exact_match_backup=True,
                                fallback_accounts=fallback_accounts, fallback_key_column=fallback_key_column, exact_match_fallback=True)
            
            # i_df, add_df concat
            i_df = pd.concat([i_df,
                            add_df], 
                            ignore_index=True
                            )

            # add_list 안의 df들에 index.name 설정
            for df in add_list :
                df_name = df.index.name
                df.index.name = f"{df_name} {year}"

            # lost_list에 하나씩 extend
            log_list.extend(add_list)

        # 이상 다른 경우 
        else:
            # 보험업일 경우
            if industry == "insurance" :
                # peroid를 "insurance"로 설정
                period = "insurance"

                # 각 기간에 해당하는 acount_list, backup_accounts, fallback_key_column 설정
                account_list = ACCOUNT_LISTS_BS[period]
                # backup_accounts = BACKUP_ACCOUNTS_BS.get(period, {}) 
                # fallback_accounts = FALLBACK_ACCOUNTS_BS.get(period, {}) 
                # fallback_key_column = "account_nm"

                # i_df 가져오기
                # add_df, add_list 설정
                add_df, add_list = extract_data(df, index_column, key_column, target_column, 
                                    account_list, exact_match_acc=True)
                
                # i_df, add_df concat
                i_df = pd.concat([i_df,
                                add_df], 
                                ignore_index=True
                                )
                
                # add_list 안의 df들에 index.name 설정
                for df in add_list :
                    df_name = df.index.name
                    df.index.name = f"{df_name} {year}"

                # lost_list에 하나씩 extend
                log_list.extend(add_list)
            
            # 2023년 이상 처리 로직
            else :
                # b기간 으로 설정
                period = "b"

                # 각 기간에 해당하는 acount_list, backup_accounts, fallback_key_column 설정
                account_list = ACCOUNT_LISTS_BS[period]
                backup_accounts = BACKUP_ACCOUNTS_BS.get(period, {}) 
                fallback_accounts = FALLBACK_ACCOUNTS_BS.get(period, {}) 
                fallback_key_column = "account_nm"

                # i_df 가져오기
                # add_df, add_list 설정
                add_df, add_list = extract_data(df, index_column, key_column, target_column, 
                                    account_list, exact_match_acc=True, 
                                    backup_accounts=backup_accounts, exact_match_backup=True,
                                    fallback_accounts=fallback_accounts, fallback_key_column=fallback_key_column, exact_match_fallback=True)
                
                # i_df, add_df concat
                i_df = pd.concat([i_df,
                                add_df], 
                                ignore_index=True
                                )
                
                # add_list 안의 df들에 index.name 설정
                for df in add_list :
                    df_name = df.index.name
                    df.index.name = f"{df_name} {year}"

                # lost_list에 하나씩 extend
                log_list.extend(add_list)
                
        i_df.rename(columns={target_column: f"{year}{suffix}"}, inplace=True)

        b_s = pd.concat([b_s, i_df], axis=1)

    b_s.set_index('account_nm', inplace=True)
    b_s.replace("len(s_row) = 0", np.nan, inplace=True)

    return b_s, log_list
################################################################
# 연간 현금흐름표 전역변수
ACCOUNT_LISTS_CF = {
    'a': [
        "ifrs_CashFlowsFromUsedInOperatingActivities",
        "ifrs_CashFlowsFromUsedInInvestingActivities", 
        "ifrs_CashFlowsFromUsedInFinancingActivities"
        ],

    'b': [
        "ifrs-full_CashFlowsFromUsedInOperatingActivities",
        "ifrs-full_CashFlowsFromUsedInInvestingActivities", 
        "ifrs-full_CashFlowsFromUsedInFinancingActivities"
	    ]
}

# 연간 현금흐름표 BACKUP 전역변수
BACKUP_ACCOUNTS_CF = {
    'a': {
        "ifrs_CashFlowsFromUsedInOperatingActivities": ["ifrs-full_CashFlowsFromUsedInOperatingActivities"],
        "ifrs_CashFlowsFromUsedInInvestingActivities": ["ifrs-full_CashFlowsFromUsedInInvestingActivities"],
        "ifrs_CashFlowsFromUsedInFinancingActivities": ["ifrs-full_CashFlowsFromUsedInFinancingActivities"]
        },

    'b': {
        # "ifrs_CashFlowsFromUsedInOperatingActivities": [""],
        # "ifrs_CashFlowsFromUsedInInvestingActivities": [""],
        # "ifrs_CashFlowsFromUsedInFinancingActivities": [""]
        }
}

# 연간 현금흐름표 FALLBACK 전역변수 
FALLBACK_ACCOUNTS_CF = {
    'a': {
        
    },

    'b': {

    }
}

# 연간 현금흐름표 parse
def parse_annual_cash_flow(fs_list):
     
    # 연간 현금흐름표 설정
    c_f = pd.DataFrame()

    # log_list 설정
    log_list = []

    for df in fs_list:

        # concat할 i_df 설정
        i_df = pd.DataFrame()

        # 연도 및 필요한 원본 데이터 설정
        year = int(df.columns.name)  
        df = df[df['sj_div'] == 'CF'][['account_id', 'account_nm', 'thstrm_amount']]

        # 연결재무제표와 별도재무제표 구분
        is_consolidated = df.index.name == "연결재무제표"
        suffix = "(연결)" if is_consolidated else "(별도)"

        # i_df의 index_column, 파싱할 key_column, 가져올 target_column 설정
        index_column = 'account_nm'
        key_column = "account_id"
        target_column = 'thstrm_amount'

        # 2015 ~ 2018년 처리 로직
        if year <= 2018:
            # a기간 으로 설정
            period = "a"            

            # 각 기간에 해당하는 acount_list, backup_accounts, fallback_key_column 설정
            account_list = ACCOUNT_LISTS_CF[period]
            backup_accounts = BACKUP_ACCOUNTS_CF.get(period, {}) 
            fallback_accounts = FALLBACK_ACCOUNTS_CF.get(period, {})
            fallback_key_column = "account_nm"

            # i_df 가져오기
            # add_df, add_list 설정
            add_df, add_list = extract_data(df, index_column, key_column, target_column, 
                                account_list, exact_match_acc=True, 
                                backup_accounts=backup_accounts, exact_match_backup=True,
                                fallback_accounts=fallback_accounts, fallback_key_column=fallback_key_column, exact_match_fallback=True)
            
            # i_df, add_df concat
            i_df = pd.concat([i_df,
                            add_df], 
                            ignore_index=True
                            )

            # add_list 안의 df들에 index.name 설정
            for df in add_list :
                df_name = df.index.name
                df.index.name = f"{df_name} {year}"

            # lost_list에 하나씩 extend
            log_list.extend(add_list)

        # 2023년 이상 처리 로직
        else :
            # b기간 으로 설정
            period = "b"

            # 각 기간에 해당하는 acount_list, backup_accounts, fallback_key_column 설정
            account_list = ACCOUNT_LISTS_CF[period]
            backup_accounts = BACKUP_ACCOUNTS_CF.get(period, {}) 
            fallback_accounts = FALLBACK_ACCOUNTS_CF.get(period, {})
            fallback_key_column = "account_nm"

            # i_df 가져오기
            # add_df, add_list 설정
            add_df, add_list = extract_data(df, index_column, key_column, target_column, 
                                account_list, exact_match_acc=True, 
                                backup_accounts=backup_accounts, exact_match_backup=True,
                                fallback_accounts=fallback_accounts, fallback_key_column=fallback_key_column, exact_match_fallback=True)
            
            # i_df, add_df concat
            i_df = pd.concat([i_df,
                            add_df], 
                            ignore_index=True
                            )
            
            # add_list 안의 df들에 index.name 설정
            for df in add_list :
                df_name = df.index.name
                df.index.name = f"{df_name} {year}"

            # lost_list에 하나씩 extend
            log_list.extend(add_list)

        i_df.rename(columns={target_column: f"{year}{suffix}"}, inplace=True)            
        c_f = pd.concat([c_f, i_df], axis=1)

    c_f.set_index('account_nm', inplace=True)
    c_f.replace("len(s_row) = 0", np.nan, inplace=True)

    return c_f, log_list
################################################################
# 연간 배당 정보
def parse_dividend(d_list):
    
    # account_lists
    account_lists = ["현금배당금총액", "주당 현금배당금"]

    # 연간 배당정보 설정
    d_s = pd.DataFrame()

    # log_list 설정
    log_list = []

    for df in d_list:

        year = df.columns.name
        df = df[['se', 'thstrm']]

        # i_df의 index_column, 파싱할 key_column, 가져올 target_column 설정
        index_column = 'se'
        key_column = "se"
        target_column = 'thstrm'

        # i_df 가져오기
        # i_df, add_list 설정
        i_df, add_list = extract_data(df, index_column, key_column, target_column, 
                            account_lists, exact_match_acc=False)

        # add_list 안의 df들에 index.name 설정
        for df in add_list :
            df_name = df.index.name
            df.index.name = f"{df_name} {year}"
        
        # lost_list에 하나씩 extend
        log_list.extend(add_list)
        
        i_df.rename(columns={target_column: f"{year}"}, inplace=True)
        d_s = pd.concat([d_s, i_df], axis=1)

    d_s.set_index('se', inplace=True)
    d_s = d_s.apply(str_to_num)

    return d_s, log_list
################################################################
# 연간 주식 수
def parse_stocks(s_list, s_name):

    # 연간 주식 수 설정
    stocks = pd.DataFrame()
    t_stocks = pd.DataFrame()

    # account_lists 설정
    account_lists = ["보통주", "우선주"]

    # log_list 설정
    log_list = []

    for df in s_list:
        
        year = df.columns.name
        df = df[["se", "istc_totqy", "tesstk_co"]]

        # i_df의 index_column, 파싱할 key_column, 가져올 target_column 설정
        index_column = 'se'
        key_column = "se"

        # 보통주 i_df 가져오기
        target_column = 'istc_totqy'

        # add_df, add_list 설정
        add_df, add_list = extract_data(df, index_column, key_column, target_column, 
                            account_lists, exact_match_acc=False)

        # i_df 저장        
        i_df = add_df

        # add_list 안의 df들에 index.name 설정
        for df in add_list :
            df_name = df.index.name
            df.index.name = f"{df_name} {year}"

        # lost_list에 하나씩 extend
        log_list.extend(add_list)

        i_df.rename(columns={'istc_totqy': year}, inplace=True)
        
        # 자사주 j_df 가져오기
        target_column = 'tesstk_co'

        # add_df, add_list 설정
        add_df, add_list = extract_data(df, index_column, key_column, target_column, 
                            account_lists, exact_match_acc=False)

        j_df = add_df

        # add_list 안의 df들에 index.name 설정
        for df in add_list :
            df_name = df.index.name
            df.index.name = f"{df_name} {year}"
        
        # lost_list에 하나씩 extend
        log_list.extend(add_list)

        j_df.rename(columns={'tesstk_co': year}, inplace=True)
        
        stocks = pd.concat([stocks, i_df[year]], axis=1)
        t_stocks = pd.concat([t_stocks, j_df[year]], axis=1)

    stocks.index = account_lists
    t_stocks.index = account_lists
    stocks.index.name = "발행한 주식 총 수"
    t_stocks.index.name = "자사주"
    stocks.columns.name = s_name
    t_stocks.columns.name = s_name

    stocks = stocks.apply(str_to_num)
    t_stocks = t_stocks.apply(str_to_num)


    return stocks, t_stocks, log_list
################################################################
# 연간 직원수 및 평균임금
def parse_labor_salary(e_list):
    lbr_sly = pd.DataFrame(index=['직원수', '평균임금'])

    for df in e_list:
        year = df.columns.name
        i_df = df[['fo_bbm', 'sm', 'fyer_salary_totamt']]

        if 'fo_bbm' in i_df.columns and i_df['fo_bbm'].notna().any():
            filt = i_df['fo_bbm'].str.contains('합계', na=False)
            filtered_df = i_df[~filt].copy()

            if not filtered_df.empty:
                filtered_df['sm'] = filtered_df['sm'].str.replace(',', '', regex=False)
                filtered_df['fyer_salary_totamt'] = filtered_df['fyer_salary_totamt'].str.replace(',', '', regex=False)

                filtered_df['sm'] = pd.to_numeric(filtered_df['sm'], errors='coerce')
                filtered_df['fyer_salary_totamt'] = pd.to_numeric(filtered_df['fyer_salary_totamt'], errors='coerce')

                total_sm = filtered_df['sm'].sum()
                total_salary = filtered_df['fyer_salary_totamt'].sum()
                avg_salary = total_salary / total_sm if total_sm != 0 else None

                result_df = pd.DataFrame({year: [total_sm, avg_salary]}, index=['직원수', '평균임금'])
                lbr_sly = pd.concat([lbr_sly, result_df], axis=1)

    return lbr_sly
################################################################
# 사업보고서 URL 
def parse_report_urls(u_df):
    report_url_df = u_df[["year", "report_url", "report_nm"]].copy()
    report_url_df.set_index("year", inplace=True)
    return report_url_df
################################################################
# 엑셀 저장하는 함수
def save_to_excel(excel_file_path, dfs_with_sheetnames, font_size, bold_font):
    with pd.ExcelWriter(excel_file_path, engine='openpyxl') as writer:
        for sheet_name, df_list in dfs_with_sheetnames:
            start_row = 0
            for df in df_list:
                df.to_excel(writer, sheet_name=sheet_name, startrow=start_row)
                start_row += len(df) + 2  # 간격

    wb = load_workbook(excel_file_path)
    for sheet_name, df_list in dfs_with_sheetnames:
        ws = wb[sheet_name]
        row = 1
        for df in df_list:
            col_len = len(df.columns) + 1  # index 포함
            # 헤더 굵게
            for col in range(1, col_len + 1):
                ws.cell(row=row, column=col).font = bold_font
            # 데이터 셀
            for r in ws.iter_rows(min_row=row + 1, max_row=row + len(df), max_col=col_len):
                for cell in r:
                    if cell.column == 1:
                        cell.font = bold_font
                    elif cell.value is not None:
                        cell.font = font_size
            row += len(df) + 2
        auto_adjust_column_width(ws)
    wb.save(excel_file_path)
################################################################
# 연구개발비
def parse_research_and_development(u_df):
    rd_tables = []

    try:
        for year, url in zip(u_df['year'][::-1], u_df['des_url_list'][::-1]):
            html_text = requests.get(url).text
            dfs = pd.read_html(StringIO(html_text))

            found = False  # 찾았는지 여부

            for df in dfs:

                # "연구개발"|"매출액 대비 비율" 포함하는 df 찾기
                if any(df.iloc[:, 0].astype(str).str.contains('연구개발비|매출액 대비 비율', na=False)):
                    df.index.name = year
                    rd_tables.append(df)
            
            # rd_tables에 값이 하나라도 있으면, found인자 True로 변경
            if bool(rd_tables) :
                found = True

            if not found:
                error_df = pd.DataFrame({"연구개발비": ["멍청해서 못찾음 ㅠ"]})
                error_df.index.name = year
                rd_tables.append(error_df)

    except Exception as e:
        # 오류 시 예외처리
        error_df = pd.DataFrame({"Exception": [f"{e}"]})
        error_df.index.name = year
        rd_tables.append(error_df)

    return rd_tables
################################################################
# 연간 자본축적 계산하는 함수 
def calculate_equity_growth_and_per(b_s, marcap, s_name):
    # 숫자 변환 및 결측값 처리
    df = b_s.apply(str_to_num)
    df.fillna(0, inplace=True)
    df = df.reset_index(drop=True)
    df = df.drop(columns=[col for col in df.columns if (df.loc[0:4, col] == 0).any()])
    
    # 7번째 행에 자본총계(지배) 계산, 인덱스 지정
    df.loc[7] = df.loc[4] - df.loc[5]
    # df = df.loc[[4,5,6,7]]
    df.index = ["유동자산","비유동자산","유동부채","비유동부채","자본총계", "비지배지분", "지배지분", "자본총계(지배)"]
    
    # 자본축적 계산
    idx = ["자본축적", "자본축적(지배)"]
    full_range = len(df.columns)

    def growth(start, end, years):
        return round((end - start) / years / 100000000)

    try:
        e_full = growth(df.iloc[4, 0], df.iloc[4, -1], full_range - 1)
        e_5y = growth(df.iloc[4, -6], df.iloc[4, -1], 5)
        e_3y = growth(df.iloc[4, -4], df.iloc[4, -1], 3)

        e_full_o = growth(df.iloc[7, 0], df.iloc[7, -1], full_range - 1)
        e_5y_o = growth(df.iloc[7, -6], df.iloc[7, -1], 5)
        e_3y_o = growth(df.iloc[7, -4], df.iloc[7, -1], 3)

    except IndexError:
        e_5y, e_5y_o = "행 모자름", "행 모자름"
        e_3y, e_3y_o = "행 모자름", "행 모자름"

    # 자본축적 DataFrame 생성
    e_df = pd.DataFrame({
        f"전체({full_range-1}년)": [e_full, e_full_o],
        "5년": [e_5y, e_5y_o],
        "3년": [e_3y, e_3y_o]
    }, index=idx)

    e_df.index.name = "(단위 : 억 원)"
    e_df.columns.name = s_name

    # PER 계산
    for label in [("자본축적", "PER"), ("자본축적(지배)", "PER(지배)")]:
        base = e_df.loc[label[0]]
        if all(isinstance(x, (int, float)) and x != 0 for x in base):
            e_df.loc[label[1]] = (marcap / base).round(1)
        else:
            e_df.loc[label[1]] = ["N/A"] * 3

    # 행 순서 정리
    desired_order = ['자본축적', 'PER', '자본축적(지배)', 'PER(지배)']
    e_per_df = e_df.reindex(desired_order)

    # 출력 확인 (선택사항, 필요시 제거 가능)
    print("=" * 30)
    print(df)
    print("=" * 30)
    print(e_per_df.to_string(float_format="%.1f"))

    return df, e_per_df
################################################################
# 연구개발 예전 코드
# def parse_research_and_development(u_df):
#     rd_df = pd.DataFrame()

#     try:
#         for year, url in zip(u_df['year'][::-1], u_df['des_url_list'][::-1]):
#             html_text = requests.get(url).text
#             dfs = pd.read_html(StringIO(html_text))

#             i = None
#             for idx, df in enumerate(dfs):
#                 if any(df.iloc[:, 0].astype(str).str.contains('연구개발')):
#                     i = idx
#                     break

#             if i is not None:
#                 df = dfs[i]
#                 filt = df.iloc[:, 0].str.contains('계|처리')
#                 i_df = df[filt].reset_index(drop=True)

#                 if i_df.iloc[0, 0] == i_df.iloc[0, 1]:
#                     s_df = i_df.iloc[:, 1:3]
#                     s_df = s_df.rename(columns={s_df.columns[1]: year})
#                 else:
#                     s_df = i_df.iloc[:, 0:2]
#                     s_df = s_df.rename(columns={s_df.columns[1]: year})

#                 rd_df = pd.concat([rd_df, s_df], axis=1)
#         print('연구개발비 파싱 완료')
#     except Exception as e:
#         print('연구개발비 table 예외처리 필요', e)

#     return rd_df